#!/usr/bin/env python3
"""Project clinical cytokine panels into DSPIN program and gene-network space."""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

try:
    from scripts import human_kg_disease, public_evidence
except ImportError:  # Supports direct execution via `python scripts/...`.
    import human_kg_disease
    import public_evidence

try:
    import openpyxl
except ImportError:  # pragma: no cover - handled at runtime when workbook support is needed.
    openpyxl = None


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DSPIN_ROOT = Path("/Users/patrickoconnell/Documents/Research Projects/10M_PBMC_project/global")
DEFAULT_PANEL_CONFIG = REPO_ROOT / "config" / "cytokine_panels" / "mayo_cypan.json"
DEFAULT_EVIDENCE_CACHE = REPO_ROOT / "artifacts" / "public_evidence_cache"
DEFAULT_KG_CACHE = REPO_ROOT / "artifacts" / "human_kg_cache"
DEFAULT_DONOR_WORKBOOK = Path(
    "/Users/patrickoconnell/Documents/Research Projects/10M_PBMC_project/"
    "A single-cell cytokine dictionary of human peripheral blood_tables.xlsx"
)
DEFAULT_CELL_METADATA = REPO_ROOT / "cell_metadata.csv"
CELL_TYPE_TOP_FRACTION = 0.01
CELL_SCORE_CHUNK_SIZE = 250_000


@dataclass
class PatientProfile:
    age: float | None
    sex: str | None
    race: str | None
    ethnicity: str | None
    cytokines: dict[str, float]


@dataclass
class AnalyteResult:
    analyte_id: str
    display_name: str
    value_pg_ml: float
    reference_upper_pg_ml: float
    log2_ratio: float
    elevation_score: float
    stimulus: str | None
    stimulus_status: str
    gene_symbols: list[str]
    note: str


def canonical_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().casefold()
    text = text.replace("α", "alpha").replace("β", "beta").replace("γ", "gamma")
    text = text.replace("_", "-").replace(" ", "-")
    text = re.sub(r"[^a-z0-9+-]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def parse_program_id(label: str) -> int | None:
    match = re.match(r"^P(\d+)", str(label))
    return int(match.group(1)) if match else None


def load_json(path: Path) -> dict[str, Any]:
    with path.open() as handle:
        return json.load(handle)


def build_alias_map(panel: dict[str, Any]) -> dict[str, dict[str, Any]]:
    alias_map: dict[str, dict[str, Any]] = {}
    for analyte in panel["analytes"]:
        names = [analyte["id"], analyte.get("display_name", analyte["id"])]
        names.extend(analyte.get("aliases", []))
        for name in names:
            alias_map[canonical_text(name)] = analyte
    return alias_map


def normalize_stimulus_name(stimulus: str, panel: dict[str, Any]) -> str:
    return panel.get("stimulus_renames", {}).get(stimulus, stimulus)


def parse_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    text = re.sub(r"^[<>]=?", "", text)
    try:
        parsed = float(text)
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def parse_cytokine_assignment(text: str) -> tuple[str, float]:
    if "=" not in text:
        raise ValueError(f"Cytokine assignment must use NAME=VALUE, got {text!r}")
    key, value = text.split("=", 1)
    parsed = parse_value(value)
    if parsed is None:
        raise ValueError(f"Could not parse cytokine value in {text!r}")
    return key.strip(), parsed


def read_patient_input(path: Path | None) -> PatientProfile:
    if path is None:
        return PatientProfile(age=None, sex=None, race=None, ethnicity=None, cytokines={})

    if path.suffix.casefold() == ".json":
        payload = load_json(path)
        cytokines = payload.get("cytokines", payload.get("panel", {}))
        if not isinstance(cytokines, dict):
            raise ValueError("JSON input must contain a cytokines object.")
        parsed = {}
        for key, value in cytokines.items():
            numeric = parse_value(value)
            if numeric is not None:
                parsed[str(key)] = numeric
        return PatientProfile(
            age=parse_value(payload.get("age")),
            sex=payload.get("sex"),
            race=payload.get("race"),
            ethnicity=payload.get("ethnicity"),
            cytokines=parsed,
        )

    with path.open(newline="") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        raise ValueError(f"No rows found in {path}")

    headers = {canonical_text(h): h for h in rows[0].keys()}
    cytokines: dict[str, float] = {}
    if "analyte" in headers and "value" in headers:
        analyte_col = headers["analyte"]
        value_col = headers["value"]
        for row in rows:
            value = parse_value(row.get(value_col))
            if value is not None:
                cytokines[str(row.get(analyte_col, "")).strip()] = value
    else:
        for header in rows[0].keys():
            value = parse_value(rows[0].get(header))
            if value is not None:
                cytokines[header] = value

    return PatientProfile(
        age=parse_value(rows[0].get(headers.get("age", ""))),
        sex=rows[0].get(headers.get("sex", "")) if "sex" in headers else None,
        race=rows[0].get(headers.get("race", "")) if "race" in headers else None,
        ethnicity=rows[0].get(headers.get("ethnicity", "")) if "ethnicity" in headers else None,
        cytokines=cytokines,
    )


def apply_profile_overrides(
    profile: PatientProfile,
    age: float | None,
    sex: str | None,
    race: str | None,
    ethnicity: str | None,
    assignments: list[str],
) -> PatientProfile:
    cytokines = dict(profile.cytokines)
    for assignment in assignments:
        key, value = parse_cytokine_assignment(assignment)
        cytokines[key] = value
    return PatientProfile(
        age=age if age is not None else profile.age,
        sex=sex if sex is not None else profile.sex,
        race=race if race is not None else profile.race,
        ethnicity=ethnicity if ethnicity is not None else profile.ethnicity,
        cytokines=cytokines,
    )


def canonicalize_cytokines(profile: PatientProfile, panel: dict[str, Any]) -> tuple[dict[str, float], list[str]]:
    alias_map = build_alias_map(panel)
    canonical: dict[str, float] = {}
    warnings: list[str] = []
    for raw_name, value in profile.cytokines.items():
        analyte = alias_map.get(canonical_text(raw_name))
        if analyte is None:
            warnings.append(f"Unrecognized cytokine analyte ignored: {raw_name}")
            continue
        canonical[analyte["id"]] = float(value)
    return canonical, warnings


def build_analyte_results(
    canonical_values: dict[str, float],
    panel: dict[str, Any],
    available_stimuli: set[str],
) -> list[AnalyteResult]:
    results = []
    for analyte in panel["analytes"]:
        analyte_id = analyte["id"]
        if analyte_id not in canonical_values:
            continue
        value = canonical_values[analyte_id]
        reference = float(analyte["reference_upper_pg_ml"])
        safe_value = max(value, 1e-9)
        safe_reference = max(reference, 1e-9)
        log2_ratio = math.log2(safe_value / safe_reference)
        elevation_score = max(0.0, log2_ratio)
        stimulus = analyte.get("stimulus")
        stimulus_status = "not_applicable"
        if stimulus:
            stimulus_status = "matched" if stimulus in available_stimuli else "missing_from_dspin"
        elif analyte.get("gene_symbols"):
            stimulus_status = "gene_seed_only"
        results.append(
            AnalyteResult(
                analyte_id=analyte_id,
                display_name=analyte.get("display_name", analyte_id),
                value_pg_ml=value,
                reference_upper_pg_ml=reference,
                log2_ratio=log2_ratio,
                elevation_score=elevation_score,
                stimulus=stimulus,
                stimulus_status=stimulus_status,
                gene_symbols=list(analyte.get("gene_symbols", [])),
                note=analyte.get("note", ""),
            )
        )
    return results


def donor_key(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = text.replace(" ", "")
    return text


def load_donor_info(workbook_path: Path, reference_year: int) -> dict[str, dict[str, Any]]:
    if not workbook_path.exists():
        return {}
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read donor workbook metadata.")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["1.donor_info"] if "1.donor_info" in workbook.sheetnames else workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    donors: dict[str, dict[str, Any]] = {}
    for row in rows:
        donor = donor_key(row[index.get("Donor number", 0)])
        if not donor:
            continue
        birth_year = parse_value(row[index.get("Birth year", -1)] if "Birth year" in index else None)
        donors[donor] = {
            "donor": donor,
            "birth_year": birth_year,
            "age": reference_year - birth_year if birth_year else None,
            "sex": row[index["Sex"]] if "Sex" in index else None,
            "race": row[index["Race"]] if "Race" in index else None,
            "ethnicity": row[index["Ethnicity"]] if "Ethnicity" in index else None,
        }
    return donors


def covariate_similarity_weights(
    donors: dict[str, dict[str, Any]],
    profile: PatientProfile,
) -> tuple[dict[str, float], list[str]]:
    if not donors:
        return {}, ["No donor metadata available; using unweighted donor averages."]

    notes: list[str] = []
    weights: dict[str, float] = {}
    for donor, info in donors.items():
        weight = 1.0
        if profile.age is not None and info.get("age") is not None:
            weight *= math.exp(-abs(float(info["age"]) - float(profile.age)) / 30.0)
        elif profile.age is not None:
            weight *= 0.8

        for key, mismatch_weight in (("sex", 0.6), ("race", 0.7), ("ethnicity", 0.7)):
            patient_value = canonical_text(getattr(profile, key))
            donor_value = canonical_text(info.get(key))
            if patient_value and donor_value:
                weight *= 1.0 if patient_value == donor_value else mismatch_weight
            elif patient_value:
                weight *= 0.85
        weights[donor] = weight

    total = sum(weights.values())
    if total <= 0:
        return {}, ["Donor covariate weights collapsed to zero; using unweighted donor averages."]
    weights = {donor: value / total for donor, value in weights.items()}
    if any(getattr(profile, key) is not None for key in ("age", "sex", "race", "ethnicity")):
        notes.append("DSPIN donor responses were covariate-weighted by available age, sex, race, and ethnicity.")
    else:
        notes.append("No patient covariates supplied; using unweighted donor averages.")
    return weights, notes


def load_response_matrix(path: Path, panel: dict[str, Any]) -> tuple[pd.DataFrame, dict[str, tuple[str, str]]]:
    frame = pd.read_csv(path, sep="\t", index_col=0)
    frame = frame.apply(pd.to_numeric, errors="coerce").fillna(0.0)
    column_meta: dict[str, tuple[str, str]] = {}
    renamed = {}
    for column in frame.columns:
        if "|" not in column:
            continue
        donor, stimulus = column.split("|", 1)
        normalized = normalize_stimulus_name(stimulus, panel)
        column_meta[column] = (donor_key(donor), normalized)
        renamed[column] = f"{donor}|{normalized}"
    if renamed:
        frame = frame.rename(columns=renamed)
        column_meta = {
            renamed.get(column, column): meta for column, meta in column_meta.items()
        }
    return frame, column_meta


def available_stimuli(column_meta: dict[str, tuple[str, str]]) -> set[str]:
    return {stimulus for _, stimulus in column_meta.values()}


def weighted_stimulus_response(
    responses: pd.DataFrame,
    column_meta: dict[str, tuple[str, str]],
    stimulus: str,
    donor_weights: dict[str, float],
) -> pd.Series | None:
    columns = [column for column, (_, stim) in column_meta.items() if stim == stimulus and column in responses.columns]
    if not columns:
        return None
    weights = np.array([donor_weights.get(column_meta[column][0], 1.0) for column in columns], dtype=float)
    if not np.isfinite(weights).all() or weights.sum() <= 0:
        weights = np.ones(len(columns), dtype=float)
    weights = weights / weights.sum()
    values = responses.loc[:, columns].to_numpy(dtype=float)
    return pd.Series(values.dot(weights), index=responses.index)


def normalize_series(series: pd.Series) -> pd.Series:
    series = series.replace([np.inf, -np.inf], np.nan).fillna(0.0).astype(float)
    max_abs = float(series.abs().max())
    if max_abs <= 0:
        return series * 0.0
    return series / max_abs


def smooth_scores(seed: pd.Series, network: pd.DataFrame | None, restart: float = 0.75) -> pd.Series:
    return normalize_series(smooth_signed_scores(seed, network, restart).clip(lower=0.0))


def smooth_signed_scores(seed: pd.Series, network: pd.DataFrame | None, restart: float = 0.75) -> pd.Series:
    seed = seed.astype(float).fillna(0.0)
    if network is None or network.empty:
        return normalize_series(seed)
    common = [name for name in seed.index if name in network.index and name in network.columns]
    if not common:
        return normalize_series(seed)
    aligned_seed = seed.loc[common]
    matrix = network.loc[common, common].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    # pandas may expose a read-only NumPy view under recent releases.
    values = matrix.to_numpy(dtype=float, copy=True)
    np.fill_diagonal(values, 0.0)
    matrix = pd.DataFrame(values, index=matrix.index, columns=matrix.columns)
    # DSPIN uses J @ state, so rows are targets and columns are sources.
    row_sum = matrix.abs().sum(axis=1).replace(0, np.nan)
    normalized = matrix.div(row_sum, axis=0).fillna(0.0)
    propagated = normalized.dot(aligned_seed)
    smoothed = restart * aligned_seed + (1.0 - restart) * propagated
    output = seed.copy()
    output.loc[common] = smoothed
    return normalize_series(output)


def load_optional_table(path: Path, index_col: int = 0) -> pd.DataFrame | None:
    if not path.exists():
        return None
    return pd.read_csv(path, sep="\t", index_col=index_col)


def build_program_scores(
    analytes: list[AnalyteResult],
    responses: pd.DataFrame,
    column_meta: dict[str, tuple[str, str]],
    donor_weights: dict[str, float],
    program_network: pd.DataFrame | None,
) -> tuple[pd.Series, pd.Series, pd.Series, list[dict[str, Any]]]:
    raw_seed = pd.Series(0.0, index=responses.index)
    total_weight = 0.0
    used: list[dict[str, Any]] = []
    for analyte in analytes:
        if analyte.elevation_score <= 0 or not analyte.stimulus:
            continue
        response = weighted_stimulus_response(responses, column_meta, analyte.stimulus, donor_weights)
        if response is None:
            continue
        weight = analyte.elevation_score
        raw_seed = raw_seed.add(response * weight, fill_value=0.0)
        total_weight += weight
        used.append(
            {
                "analyte": analyte.analyte_id,
                "stimulus": analyte.stimulus,
                "weight": weight,
            }
        )
    if total_weight > 0:
        raw_seed = raw_seed / total_weight
    seed = normalize_series(raw_seed.clip(lower=0.0))
    smoothed = smooth_scores(seed, program_network)
    signed_smoothed = smooth_signed_scores(normalize_series(raw_seed), program_network)
    return seed, smoothed, signed_smoothed, used


def build_gene_seed(analytes: list[AnalyteResult], gene_index: pd.Index) -> pd.Series:
    seed = pd.Series(0.0, index=gene_index.astype(str))
    for analyte in analytes:
        if analyte.elevation_score <= 0:
            continue
        for gene in analyte.gene_symbols:
            if gene in seed.index:
                seed.loc[gene] = max(float(seed.loc[gene]), analyte.elevation_score)
    return normalize_series(seed)


def build_gene_scores(
    program_scores: pd.Series,
    analytes: list[AnalyteResult],
    gene_to_program: pd.DataFrame | None,
    gene_network: pd.DataFrame | None,
    preserve_sign: bool = False,
) -> tuple[pd.Series, pd.Series, pd.Series]:
    if gene_to_program is None or gene_to_program.empty:
        empty = pd.Series(dtype=float)
        return empty, empty, empty

    common_programs = [name for name in gene_to_program.columns if name in program_scores.index]
    regulator_score = pd.Series(0.0, index=gene_to_program.index.astype(str))
    if common_programs:
        weights = program_scores.loc[common_programs].astype(float)
        regulator_matrix = gene_to_program.loc[:, common_programs].apply(pd.to_numeric, errors="coerce").fillna(0.0)
        regulator_score = regulator_matrix.dot(weights)
        regulator_score.index = regulator_score.index.astype(str)
        regulator_score = normalize_series(regulator_score if preserve_sign else regulator_score.clip(lower=0.0))

    direct_seed = build_gene_seed(analytes, gene_to_program.index)
    combined = normalize_series((0.75 * regulator_score).add(0.25 * direct_seed, fill_value=0.0))
    smoothed = (
        smooth_signed_scores(combined, gene_network, restart=0.8)
        if preserve_sign
        else smooth_scores(combined, gene_network, restart=0.8)
    )
    return direct_seed, regulator_score, smoothed


def load_program_annotations(path: Path, program_index: pd.Index) -> pd.DataFrame:
    if path.exists():
        frame = pd.read_csv(path, encoding="utf-8", encoding_errors="replace")
        frame["program_id"] = frame["program_id"].astype(int)
        frame["program_name"] = frame["program_label"].astype(str)
        return frame
    rows = []
    for name in program_index:
        program_id = parse_program_id(str(name))
        rows.append(
            {
                "program_id": -1 if program_id is None else program_id,
                "program_name": str(name),
                "program_label": f"P{program_id}" if program_id is not None else str(name),
                "final_annotation": str(name),
                "primary_annotation": str(name),
                "immune_cell_type": "",
                "biological_response_or_function": "",
                "confidence_tier": "",
            }
        )
    return pd.DataFrame(rows)


def module_by_node(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    frame = pd.read_csv(path, sep="\t")
    return {str(row["node"]): str(row["module"]) for _, row in frame.iterrows()}


def annotation_for_program(annotations: pd.DataFrame, program_name: str) -> dict[str, Any]:
    program_id = parse_program_id(program_name)
    if program_id is not None and "program_id" in annotations:
        hit = annotations.loc[annotations["program_id"] == program_id]
        if not hit.empty:
            return hit.iloc[0].to_dict()
    return {}


def program_scores_table(
    seed_scores: pd.Series,
    final_scores: pd.Series,
    annotations: pd.DataFrame,
    modules: dict[str, str],
) -> pd.DataFrame:
    rows = []
    for program_name, score in final_scores.sort_values(ascending=False).items():
        annotation = annotation_for_program(annotations, str(program_name))
        rows.append(
            {
                "program": program_name,
                "program_id": parse_program_id(str(program_name)),
                "score": score,
                "seed_score": float(seed_scores.get(program_name, 0.0)),
                "module": modules.get(str(program_name), ""),
                "final_annotation": annotation.get("final_annotation", annotation.get("primary_annotation", "")),
                "immune_cell_type": annotation.get("immune_cell_type", ""),
                "biological_response_or_function": annotation.get("biological_response_or_function", ""),
                "confidence_tier": annotation.get("confidence_tier", ""),
            }
        )
    return pd.DataFrame(rows)


def gene_scores_table(
    direct_seed: pd.Series,
    regulator_score: pd.Series,
    final_score: pd.Series,
    modules: dict[str, str],
    limit: int,
) -> pd.DataFrame:
    if final_score.empty:
        return pd.DataFrame(
            columns=["gene", "score", "direct_cytokine_seed", "program_regulator_score", "module"]
        )
    rows = []
    for gene, score in final_score.sort_values(ascending=False).head(limit).items():
        rows.append(
            {
                "gene": gene,
                "score": score,
                "direct_cytokine_seed": float(direct_seed.get(gene, 0.0)),
                "program_regulator_score": float(regulator_score.get(gene, 0.0)),
                "module": modules.get(str(gene), ""),
            }
        )
    return pd.DataFrame(rows)


def discover_cell_metadata_path(dspin_root: Path) -> Path | None:
    for path in [DEFAULT_CELL_METADATA, dspin_root / "cell_metadata.csv"]:
        if path.exists():
            return path
    return None


def program_annotation_cell_type_enrichment(
    program_scores: pd.Series,
    annotations: pd.DataFrame,
) -> pd.DataFrame:
    rows_by_cell_type: dict[str, dict[str, Any]] = {}
    for program_name, score in program_scores.sort_values(ascending=False).items():
        annotation = annotation_for_program(annotations, str(program_name))
        cell_type = str(annotation.get("immune_cell_type", "")).strip()
        if not cell_type:
            continue
        row = rows_by_cell_type.setdefault(
            cell_type,
            {"cell_type": cell_type, "raw_score": 0.0, "supporting_programs": []},
        )
        row["raw_score"] += max(float(score), 0.0)
        if len(row["supporting_programs"]) < 5:
            row["supporting_programs"].append(str(program_name))

    if not rows_by_cell_type:
        return pd.DataFrame(columns=["rank", "cell_type", "score", "supporting_programs", "mode"])

    table = pd.DataFrame(rows_by_cell_type.values())
    table["score"] = normalize_series(table["raw_score"])
    table["supporting_programs"] = table["supporting_programs"].apply(";".join)
    table["program_contributions"] = "{}"
    table["top_contributing_program"] = ""
    table["top_program_contribution"] = 0.0
    table["top_program_contribution_share"] = 0.0
    table["mode"] = "program_annotation_inferred"
    table = table.sort_values("score", ascending=False).reset_index(drop=True)
    table.insert(0, "rank", np.arange(1, len(table) + 1))
    return table[
        [
            "rank",
            "cell_type",
            "score",
            "supporting_programs",
            "program_contributions",
            "top_contributing_program",
            "top_program_contribution",
            "top_program_contribution_share",
            "mode",
        ]
    ]


def _read_program_names(path: Path) -> list[str]:
    frame = pd.read_csv(path, sep="\t")
    if "program" not in frame.columns:
        raise ValueError(f"{path} must contain a 'program' column.")
    return frame["program"].astype(str).tolist()


def _clean_cell_types(values: pd.Series) -> pd.Series:
    cleaned = values.fillna("Unknown").astype(str).str.strip()
    return cleaned.mask(cleaned == "", "Unknown")


def exact_cell_type_enrichment(
    program_scores: pd.Series,
    dspin_root: Path,
    metadata_path: Path,
    top_fraction: float = CELL_TYPE_TOP_FRACTION,
    chunk_size: int = CELL_SCORE_CHUNK_SIZE,
) -> pd.DataFrame:
    representation_path = dspin_root / "program_program_representation.npy"
    program_names_path = dspin_root / "program_program_names.tsv"
    obs_names_path = dspin_root / "program_obs_names.tsv"
    for path in [representation_path, program_names_path, obs_names_path]:
        if not path.exists():
            raise FileNotFoundError(path)

    metadata_columns = set(pd.read_csv(metadata_path, nrows=0).columns)
    missing_columns = {"Cell_ID", "cell_type"} - metadata_columns
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"{metadata_path} is missing required columns: {missing}")

    program_names = _read_program_names(program_names_path)
    representation = np.load(representation_path, mmap_mode="r")
    if representation.ndim != 2:
        raise ValueError(f"{representation_path} must be a 2D matrix.")
    if representation.shape[1] != len(program_names):
        raise ValueError(
            "Program matrix column count does not match program_program_names.tsv: "
            f"{representation.shape[1]} vs {len(program_names)}"
        )

    query = program_scores.reindex(program_names).fillna(0.0).astype(float).to_numpy()
    query_norm = float(np.linalg.norm(query))
    if query_norm <= 0:
        raise ValueError("Cannot compute cell-type enrichment because all patient program scores are zero.")

    n_cells = int(representation.shape[0])
    top_n = max(1, int(math.ceil(n_cells * top_fraction)))
    candidate_indices = np.array([], dtype=np.int64)
    candidate_scores = np.array([], dtype=float)
    for start in range(0, n_cells, chunk_size):
        end = min(start + chunk_size, n_cells)
        chunk = np.asarray(representation[start:end, :], dtype=float)
        denom = np.linalg.norm(chunk, axis=1) * query_norm
        scores = np.divide(chunk.dot(query), denom, out=np.zeros(end - start, dtype=float), where=denom > 0)
        take = min(top_n, scores.size)
        if take == scores.size:
            local = np.arange(scores.size)
        else:
            local = np.argpartition(scores, scores.size - take)[-take:]
        candidate_indices = np.concatenate([candidate_indices, start + local])
        candidate_scores = np.concatenate([candidate_scores, scores[local]])
        if candidate_scores.size > top_n:
            keep = np.argpartition(candidate_scores, candidate_scores.size - top_n)[-top_n:]
            candidate_indices = candidate_indices[keep]
            candidate_scores = candidate_scores[keep]

    order = np.argsort(-candidate_scores)
    top_indices = candidate_indices[order][:top_n]
    top_values = candidate_scores[order][:top_n]
    top_mask = np.zeros(n_cells, dtype=bool)
    top_scores = np.zeros(n_cells, dtype=float)
    top_mask[top_indices] = True
    top_scores[top_indices] = top_values

    total_counts: pd.Series = pd.Series(dtype=int)
    top_counts: pd.Series = pd.Series(dtype=int)
    top_score_sums: pd.Series = pd.Series(dtype=float)
    top_program_contribution_sums: dict[str, np.ndarray] = {}
    rows_seen = 0
    obs_iter = pd.read_csv(obs_names_path, sep="\t", usecols=["obs_name"], chunksize=chunk_size)
    metadata_iter = pd.read_csv(metadata_path, usecols=["Cell_ID", "cell_type"], chunksize=chunk_size)
    try:
        for metadata_chunk in metadata_iter:
            try:
                obs_chunk = next(obs_iter)
            except StopIteration as exc:
                raise ValueError(f"{metadata_path} has more rows than {obs_names_path}.") from exc
            start = rows_seen
            end = start + len(metadata_chunk)
            if end > n_cells:
                raise ValueError(f"{metadata_path} has more rows than the program matrix.")

            metadata_ids = metadata_chunk["Cell_ID"].astype(str).to_numpy()
            obs_ids = obs_chunk["obs_name"].astype(str).to_numpy()
            if len(metadata_ids) != len(obs_ids) or not np.array_equal(metadata_ids, obs_ids):
                mismatch = 0
                if len(metadata_ids) == len(obs_ids):
                    mismatch_values = np.flatnonzero(metadata_ids != obs_ids)
                    mismatch = int(mismatch_values[0]) if mismatch_values.size else 0
                row_number = start + mismatch + 1
                raise ValueError(
                    "cell_metadata.csv Cell_ID order does not match program_obs_names.tsv "
                    f"at data row {row_number}."
                )

            cell_types = _clean_cell_types(metadata_chunk["cell_type"])
            total_counts = total_counts.add(cell_types.value_counts(), fill_value=0).astype(int)
            mask = top_mask[start:end]
            if mask.any():
                top_types = cell_types.loc[mask]
                top_counts = top_counts.add(top_types.value_counts(), fill_value=0).astype(int)
                score_sums = pd.Series(top_scores[start:end][mask], index=top_types).groupby(level=0).sum()
                top_score_sums = top_score_sums.add(score_sums, fill_value=0.0)
                top_vectors = np.asarray(representation[start:end, :], dtype=float)[mask]
                contribution_denom = np.linalg.norm(top_vectors, axis=1) * query_norm
                contributions = np.divide(
                    top_vectors * query[np.newaxis, :],
                    contribution_denom[:, np.newaxis],
                    out=np.zeros_like(top_vectors, dtype=float),
                    where=contribution_denom[:, np.newaxis] > 0,
                )
                top_type_values = top_types.astype(str).to_numpy()
                for cell_type in np.unique(top_type_values):
                    contribution_sum = contributions[top_type_values == cell_type].sum(axis=0)
                    top_program_contribution_sums[cell_type] = (
                        top_program_contribution_sums.get(cell_type, np.zeros(len(program_names), dtype=float))
                        + contribution_sum
                    )
            rows_seen = end

        try:
            extra_obs = next(obs_iter)
        except StopIteration:
            extra_obs = None
    finally:
        obs_iter.close()
        metadata_iter.close()
    if extra_obs is not None:
        raise ValueError(f"{metadata_path} has fewer rows than {obs_names_path}.")
    if rows_seen != n_cells:
        raise ValueError(f"{metadata_path} row count does not match program matrix: {rows_seen} vs {n_cells}.")

    rows = []
    for cell_type, total in total_counts.items():
        top = int(top_counts.get(cell_type, 0))
        if top <= 0:
            continue
        background_fraction = float(total) / float(n_cells)
        top_cell_fraction = float(top) / float(top_n)
        fold_enrichment = top_cell_fraction / background_fraction if background_fraction > 0 else 0.0
        mean_similarity = float(top_score_sums.get(cell_type, 0.0)) / float(top)
        mean_contributions = top_program_contribution_sums.get(
            str(cell_type), np.zeros(len(program_names), dtype=float)
        ) / float(top)
        positive_contributions = np.clip(mean_contributions, 0.0, None)
        positive_total = float(positive_contributions.sum())
        top_program_index = int(np.argmax(positive_contributions)) if positive_total > 0 else None
        top_program = program_names[top_program_index] if top_program_index is not None else ""
        top_program_contribution = (
            float(positive_contributions[top_program_index]) if top_program_index is not None else 0.0
        )
        score = fold_enrichment * top_cell_fraction
        rows.append(
            {
                "cell_type": cell_type,
                "score": score,
                "fold_enrichment": fold_enrichment,
                "top_cells": top,
                "total_cells": int(total),
                "top_cell_fraction": top_cell_fraction,
                "background_fraction": background_fraction,
                "mean_similarity": mean_similarity,
                "program_contributions": json.dumps(
                    {program: float(value) for program, value in zip(program_names, mean_contributions)},
                    separators=(",", ":"),
                ),
                "top_contributing_program": top_program,
                "top_program_contribution": top_program_contribution,
                "top_program_contribution_share": (
                    top_program_contribution / positive_total if positive_total > 0 else 0.0
                ),
                "mode": "exact_per_cell",
            }
        )

    table = pd.DataFrame(rows)
    if not table.empty:
        table["score"] = normalize_series(table["score"])
    table = table.sort_values("score", ascending=False).reset_index(drop=True)
    table.insert(0, "rank", np.arange(1, len(table) + 1))
    return table


def build_cell_type_enrichment(
    program_scores: pd.Series,
    annotations: pd.DataFrame,
    dspin_root: Path,
    metadata_path: Path | None = None,
) -> tuple[pd.DataFrame, str, list[str]]:
    resolved_metadata = metadata_path if metadata_path is not None else discover_cell_metadata_path(dspin_root)
    if resolved_metadata is not None and resolved_metadata.exists():
        try:
            table = exact_cell_type_enrichment(program_scores, dspin_root, resolved_metadata)
            return table, "exact_per_cell", []
        except FileNotFoundError as exc:
            warning = (
                f"Exact per-cell cell-type enrichment unavailable because {exc.filename or exc} is missing; "
                "using program annotation-inferred cell states."
            )
    else:
        warning = "Exact per-cell cell-type enrichment unavailable because cell_metadata.csv was not found; using program annotation-inferred cell states."

    fallback = program_annotation_cell_type_enrichment(program_scores, annotations)
    return fallback, "program_annotation_inferred", [warning]


def append_hypothesis_rows(
    lines: list[str],
    table: pd.DataFrame,
    empty_message: str,
) -> None:
    if table.empty:
        lines.append(f"- {empty_message}")
        return
    for _, row in table.head(10).iterrows():
        pvalue = row.get("combined_p_value", float("nan"))
        score = row.get("evidence_score", 0.0)
        classification = row.get("hypothesis_class", "")
        pvalue_prefix = "<=" if bool(row.get("combined_p_value_is_upper_bound", False)) else "="
        permutations = row.get("direct_permutations", "")
        permutation_note = f"; permutations={permutations}" if permutations else ""
        lines.append(
            f"- {row['disease_family']}: {row['disease_name']}; score={float(score):.3f}; "
            f"p{pvalue_prefix}{float(pvalue):.3g}{permutation_note}; "
            f"{classification}; sources={row['supporting_sources']}"
        )


def append_mechanism_section(
    lines: list[str],
    title: str,
    mechanism_table: pd.DataFrame,
) -> None:
    lines.extend(["", f"## {title}", ""])
    if mechanism_table.empty:
        lines.append("- No gene-perturbation evidence matched the current cache.")
        return
    for _, row in mechanism_table.head(10).iterrows():
        lines.append(
            f"- {row['perturbation_gene']}: phenocopy={row['phenocopy_score']:.3f}; "
            f"reversal={row['reversal_score']:.3f}; sources={row['supporting_sources']}"
        )


def write_markdown_report(
    path: Path,
    profile: PatientProfile,
    analytes: list[AnalyteResult],
    program_table: pd.DataFrame,
    cell_type_table: pd.DataFrame,
    human_disease_table: pd.DataFrame,
    human_mimic_table: pd.DataFrame,
    human_mechanism_table: pd.DataFrame,
    notes: list[str],
) -> None:
    lines = [
        "# DSPIN Cytokine Panel Interpretation",
        "",
        "This is an experimental research interpretation, not a validated diagnostic report.",
        "",
        "## Patient Inputs",
        "",
        f"- Age: {profile.age if profile.age is not None else 'not provided'}",
        f"- Sex: {profile.sex or 'not provided'}",
        f"- Race: {profile.race or 'not provided'}",
        f"- Ethnicity: {profile.ethnicity or 'not provided'}",
        "",
        "## Cytokine Signals",
        "",
    ]
    for item in sorted(analytes, key=lambda x: x.elevation_score, reverse=True):
        lines.append(
            f"- {item.display_name}: {item.value_pg_ml:g} pg/mL, "
            f"log2(value/reference)={item.log2_ratio:.2f}, stimulus={item.stimulus or 'gene-only'}"
        )
    lines.extend(["", "## Top DSPIN Programs", ""])
    for _, row in program_table.head(10).iterrows():
        label = row.get("final_annotation") or row["program"]
        lines.append(f"- {row['program']}: score={row['score']:.3f}; {label}")
    lines.extend(["", "## Top Dysregulated Cell Types", ""])
    for _, row in cell_type_table.head(10).iterrows():
        if row.get("mode") == "exact_per_cell":
            lines.append(
                f"- {row['cell_type']}: score={row['score']:.3f}; "
                f"{row['fold_enrichment']:.2f}x enriched; top cells={int(row['top_cells'])}"
            )
        else:
            lines.append(
                f"- {row['cell_type']}: score={row['score']:.3f}; "
                f"programs={row.get('supporting_programs', '')}"
            )
    lines.extend(["", "## Top Disease / Mimic Hypotheses", "", "### Named Disease Evidence", ""])
    append_hypothesis_rows(
        lines,
        human_disease_table,
        "No ontology-normalized human disease evidence matched the current cache.",
    )
    lines.extend(["", "### Acquired Disease Mimics", ""])
    append_hypothesis_rows(
        lines,
        human_mimic_table,
        "No acquired mimic evidence matched the current cache.",
    )
    append_mechanism_section(lines, "Human Mechanistic Gene Perturbation Evidence", human_mechanism_table)
    lines.extend(["", "## Notes", ""])
    for note in notes:
        lines.append(f"- {note}")
    lines.append("- Gallo/JACI ratio comparator is intentionally not included.")
    path.write_text("\n".join(lines) + "\n")


def write_outputs(
    output_dir: Path,
    profile: PatientProfile,
    panel: dict[str, Any],
    analytes: list[AnalyteResult],
    program_table: pd.DataFrame,
    gene_table: pd.DataFrame,
    cell_type_table: pd.DataFrame,
    cell_type_mode: str,
    human_disease_table: pd.DataFrame,
    human_mimic_table: pd.DataFrame,
    human_mechanism_table: pd.DataFrame,
    evidence_edges: pd.DataFrame,
    network_concordance: pd.DataFrame,
    manifold_modules: pd.DataFrame,
    supporting_terms: pd.DataFrame,
    evidence_manifest: dict[str, Any],
    stimulus_uses: list[dict[str, Any]],
    donor_notes: list[str],
    warnings: list[str],
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    program_table.to_csv(output_dir / "program_scores.tsv", sep="\t", index=False)
    gene_table.to_csv(output_dir / "gene_subnetwork.tsv", sep="\t", index=False)
    cell_type_table.to_csv(output_dir / "cell_type_enrichment.tsv", sep="\t", index=False)
    human_disease_table.to_csv(output_dir / "human_disease_hypotheses.tsv", sep="\t", index=False)
    human_mimic_table.to_csv(output_dir / "human_mimic_hypotheses.tsv", sep="\t", index=False)
    human_mechanism_table.to_csv(output_dir / "human_mechanistic_perturbations.tsv", sep="\t", index=False)
    evidence_edges.to_csv(output_dir / "disease_evidence_edges.tsv", sep="\t", index=False)
    network_concordance.to_csv(output_dir / "disease_network_concordance.tsv", sep="\t", index=False)
    manifold_modules.to_csv(output_dir / "disease_manifold_modules.tsv", sep="\t", index=False)
    supporting_terms.to_csv(output_dir / "disease_supporting_terms.tsv", sep="\t", index=False)
    for filename in ["mouse_model_hypotheses.tsv", "mouse_mechanistic_perturbations.tsv"]:
        (output_dir / filename).unlink(missing_ok=True)
    (output_dir / "disease_evidence_manifest.json").write_text(json.dumps(evidence_manifest, indent=2) + "\n")

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "panel_id": panel.get("panel_id"),
        "patient": {
            "age": profile.age,
            "sex": profile.sex,
            "race": profile.race,
            "ethnicity": profile.ethnicity,
        },
        "analytes": [item.__dict__ for item in analytes],
        "stimulus_uses": stimulus_uses,
        "top_programs": program_table.head(10).to_dict(orient="records"),
        "cell_type_enrichment_mode": cell_type_mode,
        "top_cell_types": cell_type_table.head(20).to_dict(orient="records"),
        "top_genes": gene_table.head(25).to_dict(orient="records"),
        "top_human_disease_hypotheses": human_disease_table.head(20).to_dict(orient="records"),
        "top_human_mimic_hypotheses": human_mimic_table.head(20).to_dict(orient="records"),
        "top_human_mechanistic_perturbations": human_mechanism_table.head(20).to_dict(orient="records"),
        "disease_engine": "human_kg_multiview",
        "disease_evidence_manifest": evidence_manifest,
        "notes": donor_notes
        + [
            "Disease hypotheses are public-evidence experimental rankings, not diagnostic probabilities.",
            "Disease and mimic evidence is human-only; no mouse analysis or ortholog mapping is performed.",
            "Program/gene interpretation uses DSPIN outputs and available donor covariates.",
            "Cell-type interpretation uses per-cell DSPIN program similarity when cell metadata are available.",
            "Gallo/JACI ratio comparator is intentionally not included.",
        ],
        "warnings": warnings,
    }
    (output_dir / "patient_report.json").write_text(json.dumps(payload, indent=2))
    write_markdown_report(
        output_dir / "report.md",
        profile,
        analytes,
        program_table,
        cell_type_table,
        human_disease_table,
        human_mimic_table,
        human_mechanism_table,
        donor_notes + warnings,
    )


def run_interpreter(args: argparse.Namespace) -> Path:
    panel = load_json(Path(args.panel_config))
    dspin_root = Path(args.dspin_root)

    raw_profile = read_patient_input(Path(args.input) if args.input else None)
    profile = apply_profile_overrides(
        raw_profile,
        args.age,
        args.sex,
        args.race,
        args.ethnicity,
        args.cytokine,
    )
    canonical_values, warnings = canonicalize_cytokines(profile, panel)
    profile = PatientProfile(profile.age, profile.sex, profile.race, profile.ethnicity, canonical_values)
    if not profile.cytokines:
        raise ValueError("No recognized cytokine values were provided.")

    program_responses, program_column_meta = load_response_matrix(
        dspin_root / "program_relative_responses.tsv",
        panel,
    )
    program_network = load_optional_table(dspin_root / "program_network.tsv")
    gene_network = load_optional_table(dspin_root / "gene_network.tsv")
    gene_to_program = load_optional_table(dspin_root / "gene_to_program_regulators.tsv")
    annotations = load_program_annotations(dspin_root / "dspin_gene_program_annotations.csv", program_responses.index)
    program_modules = module_by_node(dspin_root / "program_modules.tsv")
    gene_modules = module_by_node(dspin_root / "gene_modules.tsv")

    donor_info = load_donor_info(
        Path(args.donor_workbook),
        int(panel.get("donor_age_reference_year", 2024)),
    )
    donor_weights, donor_notes = covariate_similarity_weights(donor_info, profile)
    analytes = build_analyte_results(profile.cytokines, panel, available_stimuli(program_column_meta))

    seed_programs, final_programs, signed_programs, stimulus_uses = build_program_scores(
        analytes,
        program_responses,
        program_column_meta,
        donor_weights,
        program_network,
    )
    direct_seed, regulator_score, final_genes = build_gene_scores(
        final_programs,
        analytes,
        gene_to_program,
        gene_network,
    )
    _, _, signed_genes = build_gene_scores(
        signed_programs,
        analytes,
        gene_to_program,
        gene_network,
        preserve_sign=True,
    )

    program_table = program_scores_table(seed_programs, final_programs, annotations, program_modules)
    gene_table = gene_scores_table(direct_seed, regulator_score, final_genes, gene_modules, args.top_genes)
    cell_type_table, cell_type_mode, cell_type_warnings = build_cell_type_enrichment(
        final_programs,
        annotations,
        dspin_root,
    )
    warnings.extend(cell_type_warnings)

    evidence_cache = Path(getattr(args, "evidence_cache", DEFAULT_EVIDENCE_CACHE))
    kg_cache = Path(getattr(args, "kg_cache", DEFAULT_KG_CACHE))
    if getattr(args, "refresh_evidence", False):
        if not getattr(args, "allow_external_patient_signature", False):
            raise ValueError(
                "--refresh-evidence sends a patient-derived signed gene list to third-party public APIs. "
                "Re-run with --allow-external-patient-signature only after confirming that disclosure is appropriate."
            )
        _, refresh_warnings = public_evidence.refresh_human_perturbseqr_cache(
            evidence_cache,
            signed_genes,
            limit=int(getattr(args, "evidence_refresh_limit", 25)),
        )
        warnings.extend(refresh_warnings)
    if getattr(args, "refresh_kg", False):
        _, kg_refresh_warnings = human_kg_disease.refresh_kg_cache(kg_cache)
        warnings.extend(kg_refresh_warnings)
    public_evidence_result = human_kg_disease.build_human_kg_hypotheses(
        kg_cache,
        evidence_cache,
        signed_genes,
        signed_programs,
        gene_network,
        gene_to_program,
        cell_type_table,
        permutations=int(getattr(args, "kg_permutations", 2_000)),
    )
    warnings.extend(public_evidence_result["warnings"])
    warnings = list(dict.fromkeys(warnings))

    output_dir = Path(args.output_dir)
    if str(output_dir) == "auto":
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = REPO_ROOT / "artifacts" / "cytokine_dspin_interpreter" / f"run_{stamp}"
    write_outputs(
        output_dir,
        profile,
        panel,
        analytes,
        program_table,
        gene_table,
        cell_type_table,
        cell_type_mode,
        public_evidence_result["human_diseases"],
        public_evidence_result["human_mimics"],
        public_evidence_result["human_mechanisms"],
        public_evidence_result["edges"],
        public_evidence_result["network_concordance"],
        public_evidence_result["manifold_modules"],
        public_evidence_result["supporting_terms"],
        public_evidence_result["manifest"],
        stimulus_uses,
        donor_notes,
        warnings,
    )
    return output_dir


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Interpret clinical cytokine panel results with DSPIN program and gene networks."
    )
    parser.add_argument("--input", default="", help="Patient JSON or CSV file.")
    parser.add_argument(
        "--cytokine",
        action="append",
        default=[],
        help="Cytokine assignment, e.g. --cytokine IL-18=1200. May be repeated.",
    )
    parser.add_argument("--age", type=float, default=None)
    parser.add_argument("--sex", default=None)
    parser.add_argument("--race", default=None)
    parser.add_argument("--ethnicity", default=None)
    parser.add_argument("--dspin-root", default=os.environ.get("DSPIN_GLOBAL_ROOT", str(DEFAULT_DSPIN_ROOT)))
    parser.add_argument("--panel-config", default=str(DEFAULT_PANEL_CONFIG))
    parser.add_argument(
        "--evidence-cache",
        default=os.environ.get("DSPIN_EVIDENCE_CACHE", str(DEFAULT_EVIDENCE_CACHE)),
        help="Local human Perturb-Seqr cache. No network access occurs unless --refresh-evidence is set.",
    )
    parser.add_argument(
        "--refresh-evidence",
        action="store_true",
        help="Refresh explicitly human Perturb-Seqr gene perturbations before scoring.",
    )
    parser.add_argument(
        "--allow-external-patient-signature",
        action="store_true",
        help="Confirm that the signed patient gene list may be sent to third-party public APIs during refresh.",
    )
    parser.add_argument(
        "--evidence-refresh-limit",
        type=int,
        default=25,
        help="Maximum human Perturb-Seqr signatures requested per signed query direction during refresh.",
    )
    parser.add_argument(
        "--kg-cache",
        default=os.environ.get("DSPIN_KG_CACHE", str(DEFAULT_KG_CACHE)),
        help="Local human ChEA-KG/Enrichr-KG cache. No network access occurs unless --refresh-kg is set.",
    )
    parser.add_argument(
        "--refresh-kg",
        action="store_true",
        help="Download static human ChEA-KG, Enrichr-KG, and MONDO assets; no patient data are transmitted.",
    )
    parser.add_argument(
        "--kg-permutations",
        type=int,
        default=2_000,
        help="Initial degree- and term-size-matched permutations for top KG terms; disease/mimic floor hits extend adaptively to 100,000.",
    )
    parser.add_argument("--donor-workbook", default=str(DEFAULT_DONOR_WORKBOOK))
    parser.add_argument("--output-dir", default="auto")
    parser.add_argument("--top-genes", type=int, default=100)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    output_dir = run_interpreter(args)
    print(f"Wrote DSPIN cytokine interpretation to {output_dir}")


if __name__ == "__main__":
    main()
